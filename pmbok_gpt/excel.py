from __future__ import annotations

from typing import Iterable, List
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet


RISK_HEADERS: List[str] = [
    "ID",
    "リスク事象",
    "カテゴリ",
    "原因",
    "影響",
    "発生確率",
    "スコア(影響×確率)",
    "対応戦略",
    "対応計画(要旨)",
    "オーナー",
    "トリガー",
    "状況",
    "メモ",
]

STAKEHOLDER_HEADERS: List[str] = [
    "ID",
    "氏名/組織",
    "役割",
    "関心事",
    "影響度(High/Med/Low)",
    "期待値",
    "関与戦略",
    "コミュニケーション(頻度/媒体)",
    "メモ",
]


def _set_headers(ws: Worksheet, headers: Iterable[str]) -> None:
    for i, h in enumerate(headers, start=1):
        ws.cell(row=1, column=i, value=h)
    ws.freeze_panes = "A2"


def create_risk_register_excel(path: str) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "RiskRegister"
    _set_headers(ws, RISK_HEADERS)

    # score = impact * probability（影響=E列, 確率=F列, スコア=G列）
    for r in range(2, 200):
        ws.cell(row=r, column=7, value=f"=E{r}*F{r}")

    wb.save(path)
    return path


def create_stakeholder_register_excel(path: str) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "Stakeholders"
    _set_headers(ws, STAKEHOLDER_HEADERS)
    wb.save(path)
    return path
